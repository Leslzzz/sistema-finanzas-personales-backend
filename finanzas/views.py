import io
import calendar
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from .models import Transaction, Budget, MonthlyPeriod


# ── Category defaults ──────────────────────────────────────────────────────────

CATEGORY_DEFAULTS = {
    'Vivienda':      {'icon': '🏠', 'color': '#60a5fa'},
    'Alimentación':  {'icon': '🍔', 'color': '#34d399'},
    'Transporte':    {'icon': '🚗', 'color': '#fbbf24'},
    'Salud':         {'icon': '💊', 'color': '#f87171'},
    'Ocio':          {'icon': '🎬', 'color': '#a78bfa'},
    'Educación':     {'icon': '📚', 'color': '#38bdf8'},
    'Ropa':          {'icon': '👕', 'color': '#fb7185'},
    'Servicios':     {'icon': '💡', 'color': '#a3e635'},
    'Ahorro':        {'icon': '💰', 'color': '#4ade80'},
    'Otros':         {'icon': '📦', 'color': '#94a3b8'},
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _period_dates(year, month):
    start = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end = date(year, month, last_day)
    return start, end


def _get_active_period(user):
    return MonthlyPeriod.objects.filter(user=user, status='active').first()


def _resolve_period(user, period_id=None):
    """Returns (period, error_response). One of them will be None."""
    if period_id:
        try:
            return MonthlyPeriod.objects.get(id=period_id, user=user), None
        except MonthlyPeriod.DoesNotExist:
            return None, Response({'message': 'Período no encontrado'}, status=404)
    period = _get_active_period(user)
    if not period:
        return None, Response(
            {'message': 'No hay un período activo. Completa el onboarding mensual.'},
            status=404,
        )
    return period, None


def _period_dict(p):
    return {
        'id': p.id,
        'year': p.year,
        'month': p.month,
        'monthlyIncome': float(p.monthly_income),
        'status': p.status,
        'startDate': p.start_date.strftime('%Y-%m-%d'),
        'endDate': p.end_date.strftime('%Y-%m-%d'),
    }


def _tx_dict(t):
    return {
        'id': str(t.id),
        'desc': t.desc,
        'amount': float(t.amount),
        'type': t.type,
        'category': t.category,
        'date': t.date.strftime('%Y-%m-%d'),
    }


def _budget_dict(b, spent):
    return {
        'id': str(b.id),
        'label': b.label,
        'icon': b.icon,
        'color': b.color,
        'limit': float(b.limit),
        'spent': float(spent),
    }


def _create_period_budgets(user, period, categories):
    for cat in categories:
        label = cat.get('label', '')
        budget_limit = cat.get('budgetLimit', 0)
        defaults = CATEGORY_DEFAULTS.get(label, {'icon': '📦', 'color': '#94a3b8'})
        Budget.objects.create(
            user=user,
            period=period,
            label=label,
            icon=defaults['icon'],
            color=defaults['color'],
            limit=Decimal(str(budget_limit)),
        )


# ── Onboarding (backward compat — delegates to period start logic) ─────────────

class OnboardingView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        today = date.today()
        year, month = today.year, today.month

        if MonthlyPeriod.objects.filter(user=user, year=year, month=month).exists():
            return Response({'message': 'Ya existe un período para este mes.'}, status=status.HTTP_409_CONFLICT)

        monthly_income = request.data.get('monthlyIncome')
        if monthly_income is not None:
            try:
                monthly_income_dec = Decimal(str(monthly_income))
            except (ValueError, InvalidOperation):
                return Response({'message': 'monthlyIncome inválido'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            monthly_income_dec = Decimal('0')

        categories = request.data.get('categories', [])

        MonthlyPeriod.objects.filter(user=user, status='active').update(
            status='closed', closed_at=timezone.now(),
        )

        start, end = _period_dates(year, month)
        period = MonthlyPeriod.objects.create(
            user=user, year=year, month=month,
            monthly_income=monthly_income_dec,
            status='active', start_date=start, end_date=end,
        )

        _create_period_budgets(user, period, categories)

        user.monthly_income = monthly_income_dec
        user.onboarding_completed = True
        user.save()

        return Response(status=status.HTTP_201_CREATED)


# ── Periods ────────────────────────────────────────────────────────────────────

class PeriodCurrentView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period = _get_active_period(request.user)
        if not period:
            return Response(
                {'message': 'No hay un período activo. Completa el onboarding mensual.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response(_period_dict(period))


class PeriodListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        periods = MonthlyPeriod.objects.filter(user=request.user)
        return Response([_period_dict(p) for p in periods])


class PeriodDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            period = MonthlyPeriod.objects.get(id=pk, user=request.user)
        except MonthlyPeriod.DoesNotExist:
            return Response({'message': 'Período no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        txs = Transaction.objects.filter(period=period)
        ingresos = txs.filter(type='ingreso').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        gastos = txs.filter(type='gasto').aggregate(total=Sum('amount'))['total'] or Decimal('0')

        data = _period_dict(period)
        data['summary'] = {
            'totalIngresos': float(ingresos),
            'totalGastos': float(gastos),
            'balance': float(ingresos - gastos),
        }
        return Response(data)


class PeriodStartView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        today = date.today()
        year, month = today.year, today.month

        if MonthlyPeriod.objects.filter(user=user, year=year, month=month).exists():
            return Response({'message': 'Ya existe un período para este mes.'}, status=status.HTTP_409_CONFLICT)

        monthly_income = request.data.get('monthlyIncome')
        if monthly_income is None:
            return Response({'message': 'monthlyIncome es requerido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            monthly_income_dec = Decimal(str(monthly_income))
            if monthly_income_dec < 0:
                raise ValueError()
        except (ValueError, InvalidOperation):
            return Response({'message': 'monthlyIncome debe ser un número positivo'}, status=status.HTTP_400_BAD_REQUEST)

        categories = request.data.get('categories', [])

        # Auto-close any active period from a previous month
        MonthlyPeriod.objects.filter(user=user, status='active').update(
            status='closed', closed_at=timezone.now(),
        )

        start, end = _period_dates(year, month)
        period = MonthlyPeriod.objects.create(
            user=user, year=year, month=month,
            monthly_income=monthly_income_dec,
            status='active', start_date=start, end_date=end,
        )

        _create_period_budgets(user, period, categories)

        user.monthly_income = monthly_income_dec
        user.onboarding_completed = True
        user.save()

        return Response(_period_dict(period), status=status.HTTP_201_CREATED)


class PeriodCloseView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        try:
            period = MonthlyPeriod.objects.get(id=pk, user=request.user)
        except MonthlyPeriod.DoesNotExist:
            return Response({'message': 'Período no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        if period.status != 'active':
            return Response({'message': 'El período ya está cerrado.'}, status=status.HTTP_400_BAD_REQUEST)

        if date.today() < period.end_date:
            return Response({'message': 'El mes actual aún no ha terminado.'}, status=status.HTTP_400_BAD_REQUEST)

        period.status = 'closed'
        period.closed_at = timezone.now()
        period.save()

        return Response({'message': 'Período cerrado correctamente.'})


# ── Transactions ───────────────────────────────────────────────────────────────

class TransactionListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period_id = request.query_params.get('periodId')
        period, err = _resolve_period(request.user, period_id)
        if err:
            return err
        txs = Transaction.objects.filter(period=period).order_by('-date')
        return Response([_tx_dict(t) for t in txs])

    def post(self, request):
        user = request.user
        period = _get_active_period(user)
        if not period:
            return Response(
                {'message': 'No tienes un período activo para este mes.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        desc = request.data.get('desc')
        amount = request.data.get('amount')
        type_ = request.data.get('type')
        category = request.data.get('category')
        date_str = request.data.get('date')

        if type_ not in ('ingreso', 'gasto'):
            return Response({'message': 'type debe ser "ingreso" o "gasto"'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                raise ValueError()
        except (ValueError, InvalidOperation):
            return Response({'message': 'amount debe ser un número positivo'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            parsed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return Response({'message': 'date debe tener formato YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)

        today = date.today()
        if parsed_date < period.start_date or parsed_date > min(period.end_date, today):
            return Response(
                {'message': f'La fecha debe estar entre {period.start_date} y {min(period.end_date, today)}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        t = Transaction.objects.create(
            user=user,
            period=period,
            desc=desc,
            amount=amount,
            type=type_,
            category=category,
            date=parsed_date,
        )
        return Response(_tx_dict(t), status=status.HTTP_201_CREATED)


class TransactionSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period_id = request.query_params.get('periodId')
        period, err = _resolve_period(request.user, period_id)
        if err:
            return err

        qs = Transaction.objects.filter(period=period)
        ingresos = qs.filter(type='ingreso').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        gastos = qs.filter(type='gasto').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        return Response({
            'ingresos': float(ingresos),
            'gastos': float(gastos),
            'balance': float(ingresos - gastos),
        })


class TransactionCategoriesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period_id = request.query_params.get('periodId')
        period, err = _resolve_period(request.user, period_id)
        if err:
            return err

        gastos = Transaction.objects.filter(period=period, type='gasto')

        totals: dict[str, Decimal] = {}
        for t in gastos:
            cat = t.category or 'Otros'
            totals[cat] = totals.get(cat, Decimal('0')) + t.amount

        if not totals:
            return Response([])

        grand_total = sum(totals.values())
        budget_colors = {b.label: b.color for b in Budget.objects.filter(period=period)}

        result = []
        for label, amount in totals.items():
            color = budget_colors.get(label) or CATEGORY_DEFAULTS.get(label, {}).get('color', '#94a3b8')
            result.append({'label': label, 'value': round(float(amount) / float(grand_total) * 100), 'color': color})

        diff = 100 - sum(r['value'] for r in result)
        if diff and result:
            result[0]['value'] += diff

        return Response(result)


class TransactionImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        import unicodedata
        import pandas as pd

        user = request.user
        period = _get_active_period(user)
        if not period:
            return Response(
                {'message': 'No tienes un período activo para este mes.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file = request.FILES.get('file')
        if not file:
            return Response({'message': 'No se proporcionó archivo'}, status=status.HTTP_400_BAD_REQUEST)

        name = file.name.lower()
        try:
            if name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file, dtype=str)
            elif name.endswith('.csv'):
                df = pd.read_csv(file, dtype=str, encoding='utf-8-sig')
            else:
                return Response({'message': 'Formato no soportado. Use .csv o .xlsx'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'message': f'Error al leer el archivo: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        def normalize(s):
            return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode().lower().strip()

        df.columns = [normalize(c) for c in df.columns]

        col_map = {
            'date':     ['fecha', 'date'],
            'desc':     ['descripcion', 'desc', 'description', 'concepto', 'detalle'],
            'amount':   ['monto', 'amount', 'importe', 'valor'],
            'type':     ['tipo', 'type'],
            'category': ['categoria', 'category', 'cat'],
        }

        def find_col(field):
            for alias in col_map[field]:
                if alias in df.columns:
                    return alias
            return None

        col_date     = find_col('date')
        col_desc     = find_col('desc')
        col_amount   = find_col('amount')
        col_type     = find_col('type')
        col_category = find_col('category')

        if not col_desc or not col_amount:
            return Response({
                'message': 'Columnas no reconocidas. El archivo debe tener: Fecha, Descripción, Monto, Tipo, Categoría'
            }, status=status.HTTP_400_BAD_REQUEST)

        imported = 0
        skipped = 0
        errors = []
        to_create = []

        for idx, row in df.iterrows():
            row_num = idx + 2

            desc = str(row.get(col_desc, '')).strip()
            if not desc or desc == 'nan':
                errors.append({'row': row_num, 'msg': 'Descripción vacía'})
                skipped += 1
                continue

            try:
                amount_raw = str(row.get(col_amount, '')).replace('$', '').replace(',', '').strip()
                amount = Decimal(amount_raw)
                if amount <= 0:
                    raise ValueError()
            except (ValueError, InvalidOperation):
                errors.append({'row': row_num, 'msg': 'Monto inválido'})
                skipped += 1
                continue

            raw_type = str(row.get(col_type, 'gasto')).lower().strip() if col_type else 'gasto'
            tx_type = 'ingreso' if raw_type.startswith('i') else 'gasto'

            category = str(row.get(col_category, 'Otros')).strip() if col_category else 'Otros'
            if not category or category == 'nan':
                category = 'Otros'

            tx_date = date.today()
            if col_date:
                try:
                    tx_date = pd.to_datetime(str(row.get(col_date, '')), dayfirst=False).date()
                except Exception:
                    pass

            if tx_date < period.start_date or tx_date > period.end_date:
                errors.append({'row': row_num, 'msg': f'Fecha {tx_date} fuera del rango del período activo'})
                skipped += 1
                continue

            to_create.append(Transaction(
                user=user,
                period=period,
                desc=desc,
                amount=amount,
                type=tx_type,
                category=category,
                date=tx_date,
            ))
            imported += 1

        if to_create:
            Transaction.objects.bulk_create(to_create)

        return Response({'imported': imported, 'skipped': skipped, 'errors': errors})


class TransactionTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, _request):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Plantilla'

        ws.append(['Fecha', 'Descripción', 'Monto', 'Tipo', 'Categoría'])

        ref = ['Categorías válidas:', 'Tipo: ingreso | gasto', 'Fecha: YYYY-MM-DD', ''] + list(CATEGORY_DEFAULTS.keys())
        for i, line in enumerate(ref, start=1):
            ws.cell(row=i, column=7, value=line)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="finanzly-plantilla.xlsx"'
        return resp


class TransactionExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        fmt = request.query_params.get('format', 'csv').lower()
        transactions = Transaction.objects.filter(user=request.user).order_by('-date')

        if fmt == 'csv':
            return self._export_csv(transactions)
        elif fmt == 'pdf':
            return self._export_pdf(transactions, request.user)
        return Response({'message': 'Formato no soportado. Usa ?format=csv o ?format=pdf'}, status=status.HTTP_400_BAD_REQUEST)

    def _export_csv(self, transactions):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transacciones'

        headers = ['Fecha', 'Descripción', 'Monto', 'Tipo', 'Categoría']
        ws.append(headers)

        for t in transactions:
            ws.append([
                t.date.strftime('%Y-%m-%d') if t.date else '',
                t.desc,
                float(t.amount),
                t.type,
                t.category or 'Otros',
            ])

        ref = ['Categorías válidas:', 'Tipo: ingreso | gasto', ''] + list(CATEGORY_DEFAULTS.keys())
        for i, line in enumerate(ref, start=1):
            ws.cell(row=i, column=7, value=line)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        resp = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="finanzly-transacciones.xlsx"'
        return resp

    def _export_pdf(self, transactions, user):
        from weasyprint import HTML
        from django.template.loader import render_to_string

        ingresos = sum(t.amount for t in transactions if t.type == 'ingreso')
        gastos = sum(t.amount for t in transactions if t.type == 'gasto')
        balance = ingresos - gastos

        html_string = render_to_string('transactions/export_pdf.html', {
            'transactions': transactions,
            'summary': {'ingresos': ingresos, 'gastos': gastos, 'balance': balance},
            'user': user,
            'today': date.today().strftime('%d de %B de %Y'),
        })

        pdf_bytes = HTML(string=html_string).write_pdf()
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="finanzly-reporte.pdf"'
        return resp


# ── Budgets ────────────────────────────────────────────────────────────────────

class BudgetListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period_id = request.query_params.get('periodId')
        period, err = _resolve_period(request.user, period_id)
        if err:
            return err

        budgets = Budget.objects.filter(period=period)

        result = []
        for b in budgets:
            spent = Transaction.objects.filter(
                period=period, type='gasto', category=b.label,
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
            result.append(_budget_dict(b, spent))
        return Response(result)


class BudgetDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        user = request.user
        try:
            budget = Budget.objects.get(id=pk, user=user)
        except Budget.DoesNotExist:
            return Response({'message': 'Presupuesto no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        active_period = _get_active_period(user)
        if not active_period or budget.period_id != active_period.id:
            return Response(
                {'message': 'Solo puedes editar presupuestos del período activo.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        new_limit = request.data.get('limit')
        if new_limit is not None:
            budget.limit = Decimal(str(new_limit))
            budget.save()

        spent = Transaction.objects.filter(
            period=active_period, type='gasto', category=budget.label,
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        return Response(_budget_dict(budget, spent))
